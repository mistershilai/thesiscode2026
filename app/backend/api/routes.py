"""FastAPI route definitions."""

import io
import os
from pathlib import Path

import numpy as np
import pandas as pd
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse

import re

from ..core.data_loader import (
    app_data,
    BASE_DIR,
    PIPELINE_DIR,
    CMS_SCENARIOS_DIR,
    BUILTIN_SCENARIO_IDS,
)
from ..core.optimizer import (
    build_region_instance,
    build_cms_region_instance,
    nb_sigma_from_mean,
    run_simulation,
    run_planning,
)
from .schemas import OptimizationRequest, OptimizationResult, PlanningRequest

router = APIRouter(prefix="/api")


@router.get("/health")
def health_check():
    return {"status": "ok", "data_loaded": app_data.loaded}


@router.get("/summary")
def get_summary():
    return app_data.get_facility_summary()


@router.get("/regions")
def list_regions():
    regions = []
    for dhmt in app_data.dhmt_list:
        if dhmt.strip() in ("", "--"):
            continue
        fac_count = len(app_data.get_facilities_for_region(dhmt))
        source = app_data.DHMT_SOURCE_MAP.get(dhmt, app_data.CMS_NAME)
        regions.append({
            "name": dhmt,
            "facility_count": fac_count,
            "source_node": source,
        })
    return regions


@router.get("/regions/{region}/facilities")
def get_region_facilities(region: str):
    fac = app_data.get_facilities_for_region(region)
    if fac.empty:
        raise HTTPException(404, f"No facilities for region: {region}")
    records = []
    for _, row in fac.iterrows():
        records.append({
            "name": str(row.get("Facility Name", "")),
            "type": str(row.get("Service Delivery Type", "")),
            "latitude": float(row["latitude"]),
            "longitude": float(row["longitude"]),
            "dhmt": str(row.get("DHMT", "")),
        })
    return records


@router.get("/facilities/geojson")
def get_facilities_geojson():
    return app_data.get_all_facilities_geojson()


@router.get("/districts/geojson")
def get_districts_geojson():
    return app_data.get_districts_geojson()


@router.get("/cms/products")
def get_cms_products():
    cms = app_data.cms_active
    scenario_ids = [s["id"] for s in app_data.list_scenarios()]
    records = []
    for code, row in cms.iterrows():
        biweekly = {
            sid: float(row.get(f"biweekly_{sid}", 0) or 0) for sid in scenario_ids
        }
        records.append({
            "product_code": str(code),
            "description": str(row.get("description", "")),
            "unit_price_bwp": float(row.get("unit_price_bwp", 0)),
            # kept for backward compatibility
            "biweekly_2526": biweekly.get("2526", 0.0),
            "biweekly_2627": biweekly.get("2627", 0.0),
            "biweekly": biweekly,
        })
    return records


@router.post("/cms/products/add")
def add_cms_product(body: dict):
    """Add a new drug product and save to disk."""
    code = body.get("product_code", "").strip()
    desc = body.get("description", "").strip()
    price = float(body.get("unit_price_bwp", 0))
    biweekly = float(body.get("biweekly_demand", 0))
    if not code:
        raise HTTPException(400, "product_code is required")

    cms = app_data.cms_active
    if code in cms.index:
        raise HTTPException(400, f"Product {code} already exists")

    new_vals = {"description": desc, "unit_price_bwp": price}
    # seed the new product's demand into every scenario column that exists
    for col in cms.columns:
        if col.startswith("biweekly_"):
            new_vals[col] = biweekly
    new_row = pd.Series(new_vals, name=code)
    app_data.cms_active = pd.concat([cms, new_row.to_frame().T])
    app_data.cms_proc_cost = app_data.cms_active["unit_price_bwp"]
    _save_cms()
    for name in getattr(app_data, "custom_scenarios", []):
        _save_scenario(name)
    return {"status": "ok", "product_code": code, "total_products": len(app_data.cms_active)}


@router.post("/cms/products/remove")
def remove_cms_product(body: dict):
    """Remove a drug product and save to disk."""
    code = body.get("product_code", "").strip()
    if not code:
        raise HTTPException(400, "product_code is required")

    cms = app_data.cms_active
    if code not in cms.index:
        raise HTTPException(404, f"Product {code} not found")

    desc = str(cms.loc[code].get("description", code))
    app_data.cms_active = cms.drop(index=code)
    app_data.cms_proc_cost = app_data.cms_active["unit_price_bwp"]
    _save_cms()
    for name in getattr(app_data, "custom_scenarios", []):
        _save_scenario(name)
    return {"status": "ok", "removed": code, "description": desc, "total_products": len(app_data.cms_active)}


# CMS demand scenarios (extra biweekly_<name> columns, persisted under cms_scenarios/)

_SCENARIO_NAME_RE = re.compile(r"^[A-Za-z0-9_-]{1,40}$")


def _save_scenario(name: str):
    """Write a custom scenario's per-product biweekly demand to its own CSV."""
    col = f"biweekly_{name}"
    CMS_SCENARIOS_DIR.mkdir(parents=True, exist_ok=True)
    df = app_data.cms_active[[col]].rename(columns={col: "biweekly"})
    df.index.name = "product_code"
    df.to_csv(CMS_SCENARIOS_DIR / f"{name}.csv")


@router.get("/cms/scenarios")
def list_cms_scenarios():
    return app_data.list_scenarios()


@router.post("/cms/scenarios/add")
def add_cms_scenario(body: dict):
    """Create a new demand scenario by duplicating an existing one."""
    name = str(body.get("name", "")).strip()
    copy_from = str(body.get("copy_from", "")).strip()
    if not _SCENARIO_NAME_RE.match(name):
        raise HTTPException(400, "Name must be 1-40 chars: letters, numbers, '_' or '-'")
    if name in BUILTIN_SCENARIO_IDS:
        raise HTTPException(400, f"'{name}' is a reserved built-in scenario id")
    if app_data.scenario_exists(name):
        raise HTTPException(400, f"Scenario '{name}' already exists")
    if not app_data.scenario_exists(copy_from):
        raise HTTPException(400, f"Source scenario '{copy_from}' not found")

    app_data.cms_active[f"biweekly_{name}"] = app_data.cms_active[f"biweekly_{copy_from}"].copy()
    app_data.custom_scenarios.append(name)
    _save_scenario(name)
    return {"status": "ok", "scenarios": app_data.list_scenarios()}


@router.post("/cms/scenarios/update")
def update_cms_scenario(body: dict):
    """Update per-product biweekly demand values for a custom scenario."""
    name = str(body.get("name", "")).strip()
    values = body.get("values", {}) or {}
    if name in BUILTIN_SCENARIO_IDS:
        raise HTTPException(400, "Built-in scenarios cannot be edited")
    if not app_data.scenario_exists(name):
        raise HTTPException(404, f"Scenario '{name}' not found")

    col = f"biweekly_{name}"
    updated = 0
    for code, val in values.items():
        if code in app_data.cms_active.index:
            try:
                app_data.cms_active.loc[code, col] = float(val)
                updated += 1
            except (TypeError, ValueError):
                continue
    _save_scenario(name)
    return {"status": "ok", "updated": updated, "scenarios": app_data.list_scenarios()}


@router.post("/cms/scenarios/remove")
def remove_cms_scenario(body: dict):
    """Delete a custom scenario (built-ins cannot be removed)."""
    name = str(body.get("name", "")).strip()
    if name in BUILTIN_SCENARIO_IDS:
        raise HTTPException(400, "Built-in scenarios cannot be removed")
    if name not in getattr(app_data, "custom_scenarios", []):
        raise HTTPException(404, f"Scenario '{name}' not found")

    app_data.cms_active = app_data.cms_active.drop(columns=[f"biweekly_{name}"], errors="ignore")
    app_data.custom_scenarios.remove(name)
    path = CMS_SCENARIOS_DIR / f"{name}.csv"
    if path.exists():
        path.unlink()
    return {"status": "ok", "removed": name, "scenarios": app_data.list_scenarios()}


OSRM_URL = os.environ.get("OSRM_URL", "http://localhost:5001")

# Persist helpers

FAC_CSV = BASE_DIR / "data/processed/facilities_with_warehouses.csv"
CMS_CSV = PIPELINE_DIR / "antimicrobials.csv"
DIST_CSV = BASE_DIR / "data/processed/distance_matrix_named.csv"
DUR_CSV = BASE_DIR / "data/processed/duration_matrix_named.csv"


def _save_facilities():
    app_data.fac.to_csv(FAC_CSV, index=False)


def _save_cms():
    # write only the core columns so the file round-trips through _load_cms_data;
    # custom scenario columns are persisted separately under cms_scenarios/
    core = ["description", "unit_price_bwp", "biweekly_2526", "biweekly_2627"]
    df = app_data.cms_active[[c for c in core if c in app_data.cms_active.columns]].copy()
    df.index.name = "product_code"
    df.to_csv(CMS_CSV)


def _save_matrices():
    app_data.dist_matrix_df.to_csv(DIST_CSV)
    app_data.time_matrix_df.to_csv(DUR_CSV)


OSRM_PORT = int(os.environ.get("OSRM_PORT", "5001"))
OSRM_IMAGE = "ghcr.io/project-osrm/osrm-backend:v5.27.1"
OSRM_CONTAINER = "kaelo-osrm"
OSRM_DATA_DIR = str(
    (Path(__file__).resolve().parent.parent.parent.parent / "osrm_project").resolve()
)


def _ensure_osrm():
    """
    Ensure OSRM is reachable. If running inside Docker Compose, the osrm
    service is managed externally. If running locally, auto-start a container.
    """
    import subprocess, time, requests as _req

    # Quick check: is it already responding?
    try:
        r = _req.get(f"{OSRM_URL}/nearest/v1/driving/25.9,-24.6", timeout=3)
        if r.status_code == 200:
            return True
    except Exception:
        pass

    # If OSRM_URL points to a Docker Compose service (not localhost),
    # OSRM cannot be auto-started here; it should already be running
    if "localhost" not in OSRM_URL and "127.0.0.1" not in OSRM_URL:
        return False

    # Check if Docker is available
    try:
        subprocess.run(["docker", "info"], capture_output=True, check=True, timeout=5)
    except Exception:
        return False

    # Remove stale container with same name (if stopped)
    subprocess.run(
        ["docker", "rm", "-f", OSRM_CONTAINER],
        capture_output=True, timeout=10,
    )

    # Start container
    subprocess.run([
        "docker", "run", "-d",
        "--name", OSRM_CONTAINER,
        "-p", f"{OSRM_PORT}:5000",
        "-v", f"{OSRM_DATA_DIR}:/data",
        OSRM_IMAGE,
        "osrm-routed", "--algorithm", "mld", "/data/botswana-latest.osrm",
    ], capture_output=True, timeout=30)

    # Wait for it to be ready (up to 15s)
    for _ in range(15):
        time.sleep(1)
        try:
            r = _req.get(f"{OSRM_URL}/nearest/v1/driving/25.9,-24.6", timeout=2)
            if r.status_code == 200:
                return True
        except Exception:
            continue
    return False


@router.post("/facilities/add")
def add_facility(body: dict):
    """
    Add a new facility and compute distances to existing facilities.
    Tries OSRM for road distances; falls back to haversine (crow-fly).
    """
    import requests as _requests

    name = body.get("name", "").strip()
    fac_type = body.get("type", "Clinic").strip()
    dhmt = body.get("dhmt", "").strip()
    lat = float(body.get("latitude", 0))
    lon = float(body.get("longitude", 0))
    parent_hospital = body.get("parent_hospital", "").strip()

    if not name or not dhmt:
        raise HTTPException(400, "name and dhmt are required")
    if lat == 0 or lon == 0:
        raise HTTPException(400, "Valid latitude and longitude required")

    existing = app_data.fac["Facility Name"].astype(str).str.strip()
    if name in existing.values:
        raise HTTPException(400, f"Facility '{name}' already exists")

    new_row = {
        "Facility Name": name,
        "Service Delivery Type": fac_type,
        "DHMT": dhmt,
        "latitude": lat,
        "longitude": lon,
        "Is_Warehouse": False,
        "Facility Status": "Active",
        "parent_hospital": parent_hospital or None,
    }
    app_data.fac = pd.concat([app_data.fac, pd.DataFrame([new_row])], ignore_index=True)

    # Gather coordinates for existing nodes
    existing_nodes = list(app_data.dist_matrix_df.index)
    fac_coords = {}
    for _, row in app_data.fac.iterrows():
        fn = str(row.get("Facility Name", "")).strip()
        if fn in existing_nodes:
            fac_coords[fn] = (float(row["longitude"]), float(row["latitude"]))

    # Ensure OSRM is running; auto-start Docker container if needed
    if not _ensure_osrm():
        raise HTTPException(
            503,
            f"Could not start OSRM routing server. "
            "Road distances are required to add facilities. "
            "Ensure Docker is installed and the OSRM data is at "
            f"{OSRM_DATA_DIR}/botswana-latest.osrm"
        )

    new_dists = {}
    new_times = {}
    for node, (nlon, nlat) in fac_coords.items():
        try:
            r = _requests.get(
                f"{OSRM_URL}/route/v1/driving/{lon},{lat};{nlon},{nlat}",
                params={"overview": "false"},
                timeout=5,
            )
            data = r.json()
            if data.get("code") == "Ok":
                route = data["routes"][0]
                new_dists[node] = route["distance"]   # meters
                new_times[node] = route["duration"]    # seconds
            else:
                new_dists[node] = float("inf")
                new_times[node] = float("inf")
        except Exception:
            new_dists[node] = float("inf")
            new_times[node] = float("inf")

    # Expand distance matrix
    dm = app_data.dist_matrix_df
    dm.loc[name] = pd.Series({n: new_dists.get(n, float("inf")) for n in dm.columns})
    dm[name] = pd.Series({n: new_dists.get(n, float("inf")) for n in dm.index})
    dm.loc[name, name] = 0.0
    app_data.dist_matrix_df = dm

    # Expand time matrix
    tm = app_data.time_matrix_df
    tm.loc[name] = pd.Series({n: new_times.get(n, float("inf")) for n in tm.columns})
    tm[name] = pd.Series({n: new_times.get(n, float("inf")) for n in tm.index})
    tm.loc[name, name] = 0.0
    app_data.time_matrix_df = tm

    computed = sum(1 for v in new_dists.values() if v != float("inf"))
    _save_facilities()
    _save_matrices()
    return {
        "status": "ok",
        "facility": name,
        "distances_computed": computed,
        "total_facilities": len(app_data.fac),
        "distance_method": "osrm",
    }


@router.post("/facilities/remove")
def remove_facility(body: dict):
    """
    Remove a facility with network-aware validation:
    - Warehouses: BLOCKED (every region needs one)
    - Hospitals: requires 'replacement_hospital' in body
    - Clinics/Health Posts: removed freely
    """
    name = body.get("name", "").strip()
    if not name:
        raise HTTPException(400, "name is required")

    fac = app_data.fac
    match = fac[fac["Facility Name"].astype(str).str.strip() == name]
    if match.empty:
        raise HTTPException(404, f"Facility '{name}' not found")

    fac_type = str(match.iloc[0].get("Service Delivery Type", "")).strip()
    dhmt = str(match.iloc[0].get("DHMT", "")).strip()

    # Warehouses: never remove
    warehouse_types = {"Warehouse"}
    if fac_type in warehouse_types or bool(match.iloc[0].get("Is_Warehouse", False)):
        raise HTTPException(
            400,
            f"Cannot remove warehouse '{name}'. "
            "Every region requires a warehouse as its supply source."
        )

    # Hospitals: require replacement
    hospital_types = {"Primary Hospital", "District Hospital", "Referral Hospital"}
    if fac_type in hospital_types:
        replacement = body.get("replacement_hospital", "").strip()
        if not replacement:
            # Return info about downstream facilities so the UI can prompt
            region_fac = fac[fac["DHMT"].astype(str).str.strip() == dhmt]
            downstream = region_fac[
                region_fac["Service Delivery Type"].isin(["Clinic", "Clinic with Maternity", "Health Post"])
            ]["Facility Name"].astype(str).str.strip().tolist()
            other_hospitals = region_fac[
                (region_fac["Service Delivery Type"].isin(hospital_types))
                & (region_fac["Facility Name"].astype(str).str.strip() != name)
            ]["Facility Name"].astype(str).str.strip().tolist()

            raise HTTPException(
                409,  # Conflict: needs user action
                detail={
                    "message": f"Removing hospital '{name}' requires rerouting downstream facilities.",
                    "downstream_facilities": downstream,
                    "available_hospitals": other_hospitals,
                    "action": "Provide 'replacement_hospital' (existing name or add a new hospital first).",
                },
            )

        # Validate replacement exists and is a hospital
        repl_match = fac[fac["Facility Name"].astype(str).str.strip() == replacement]
        if repl_match.empty:
            raise HTTPException(400, f"Replacement hospital '{replacement}' not found")
        repl_type = str(repl_match.iloc[0].get("Service Delivery Type", ""))
        if repl_type not in hospital_types:
            raise HTTPException(400, f"'{replacement}' is a {repl_type}, not a hospital")

    # Remove facility
    app_data.fac = fac[fac["Facility Name"].astype(str).str.strip() != name].reset_index(drop=True)

    # Remove from distance/time matrices
    for matrix in (app_data.dist_matrix_df, app_data.time_matrix_df):
        if name in matrix.index:
            matrix.drop(index=name, inplace=True)
        if name in matrix.columns:
            matrix.drop(columns=name, inplace=True)

    _save_facilities()
    _save_matrices()
    return {
        "status": "ok",
        "removed": name,
        "type": fac_type,
        "total_facilities": len(app_data.fac),
    }


@router.post("/facilities/relocate")
def relocate_facility(body: dict):
    """
    Update a facility's coordinates and recompute OSRM distances.
    Use this to move a warehouse, hospital, or any facility to a new location.
    """
    import requests as _requests

    name = body.get("name", "").strip()
    new_lat = body.get("latitude")
    new_lon = body.get("longitude")

    if not name:
        raise HTTPException(400, "name is required")
    if new_lat is None or new_lon is None:
        raise HTTPException(400, "latitude and longitude are required")

    new_lat, new_lon = float(new_lat), float(new_lon)

    fac = app_data.fac
    mask = fac["Facility Name"].astype(str).str.strip() == name
    if not mask.any():
        raise HTTPException(404, f"Facility '{name}' not found")

    # Update coordinates
    app_data.fac.loc[mask, "latitude"] = new_lat
    app_data.fac.loc[mask, "longitude"] = new_lon

    # Ensure OSRM is available
    if not _ensure_osrm():
        raise HTTPException(503, "OSRM not available; cannot recompute distances.")

    # Gather coordinates for all other nodes in the distance matrix
    existing_nodes = [n for n in app_data.dist_matrix_df.index if n != name]
    fac_coords = {}
    for _, row in app_data.fac.iterrows():
        fn = str(row.get("Facility Name", "")).strip()
        if fn in existing_nodes:
            fac_coords[fn] = (float(row["longitude"]), float(row["latitude"]))

    # Recompute distances from this facility to all others
    new_dists = {}
    new_times = {}
    for node, (nlon, nlat) in fac_coords.items():
        try:
            r = _requests.get(
                f"{OSRM_URL}/route/v1/driving/{new_lon},{new_lat};{nlon},{nlat}",
                params={"overview": "false"},
                timeout=5,
            )
            data = r.json()
            if data.get("code") == "Ok":
                route = data["routes"][0]
                new_dists[node] = route["distance"]
                new_times[node] = route["duration"]
            else:
                new_dists[node] = float("inf")
                new_times[node] = float("inf")
        except Exception:
            new_dists[node] = float("inf")
            new_times[node] = float("inf")

    # Update distance matrix
    dm = app_data.dist_matrix_df
    if name not in dm.index:
        dm.loc[name] = float("inf")
        dm[name] = float("inf")
    for node, dist in new_dists.items():
        dm.loc[name, node] = dist
        dm.loc[node, name] = dist
    dm.loc[name, name] = 0.0

    # Update time matrix
    tm = app_data.time_matrix_df
    if name not in tm.index:
        tm.loc[name] = float("inf")
        tm[name] = float("inf")
    for node, t in new_times.items():
        tm.loc[name, node] = t
        tm.loc[node, name] = t
    tm.loc[name, name] = 0.0

    computed = sum(1 for v in new_dists.values() if v != float("inf"))
    _save_facilities()
    _save_matrices()
    return {
        "status": "ok",
        "facility": name,
        "new_coordinates": {"latitude": new_lat, "longitude": new_lon},
        "distances_recomputed": computed,
    }


@router.get("/regions/{region}/demand")
def get_region_demand(region: str, scenario: str = "2526", use_cms: bool = True):
    """
    Return the default computed demand matrix for a region.
    Response: { facilities: [...], drug_classes: [...], demand: {facility: {drug: value}} }
    Users can edit these values and send them back via custom_demand in /optimize.
    """
    try:
        if use_cms:
            instance = build_cms_region_instance(region, scenario=scenario)
        else:
            instance = build_region_instance(region)
    except ValueError as e:
        raise HTTPException(400, str(e))

    mu_mat = instance["mu_mat"]
    demand_dict = {}
    for fac in mu_mat.index:
        row = mu_mat.loc[fac]
        demand_dict[fac] = {
            col: round(float(val), 2) for col, val in row.items() if val > 0
        }

    return {
        "region": region,
        "facilities": list(mu_mat.index),
        "drug_classes": list(mu_mat.columns),
        "demand": demand_dict,
    }


@router.post("/optimize", response_model=OptimizationResult)
def run_optimization(req: OptimizationRequest):
    try:
        if req.use_cms_data:
            instance = build_cms_region_instance(req.region, scenario=req.scenario)
        else:
            instance = build_region_instance(req.region)
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Apply demand multiplier
    if req.demand_multiplier != 1.0:
        instance["mu_mat"] = instance["mu_mat"] * req.demand_multiplier
        instance["sigma_mat"] = nb_sigma_from_mean(instance["mu_mat"], kappa=req.kappa)

    # Apply custom demand overrides
    if req.custom_demand:
        mu_mat = instance["mu_mat"].copy()
        for fac_name, drug_demands in req.custom_demand.items():
            if fac_name not in mu_mat.index:
                continue
            for drug_class, value in drug_demands.items():
                if drug_class in mu_mat.columns:
                    mu_mat.loc[fac_name, drug_class] = float(value)
        instance["mu_mat"] = mu_mat
        instance["sigma_mat"] = nb_sigma_from_mean(mu_mat, kappa=req.kappa)

    # Shortage penalty = multiplier × per-drug procurement cost
    proc_cost = instance.get("proc_cost", 0.0)
    if np.isscalar(proc_cost):
        shortage_pen = req.shortage_penalty * float(proc_cost) if proc_cost else 10.0
    else:
        shortage_pen = pd.Series(proc_cost).clip(lower=1.0) * req.shortage_penalty

    try:
        metrics_df = run_simulation(
            instance=instance,
            strategy=req.strategy,
            T=req.periods,
            kappa=req.kappa,
            Gamma=req.gamma,
            transport_cost_per_km=req.transport_cost_per_km,
            shortage_penalty=shortage_pen,
            holding_cost=req.holding_cost,
            procurement_cost=proc_cost,
            supply_multiplier=req.supply_multiplier,
            seed=req.seed,
        )
    except Exception as e:
        raise HTTPException(500, f"Optimization failed: {e}")

    periods = metrics_df.to_dict(orient="records")

    # Compute summary
    valid = metrics_df[metrics_df["status"] == "optimal"]
    summary = {}
    if not valid.empty:
        summary = {
            "avg_unmet_pct": round(float(valid["unmet_pct"].mean()), 2),
            "max_unmet_pct": round(float(valid["unmet_pct"].max()), 2),
            "total_cost": round(float(valid["objective"].sum()), 2),
            "total_transport_cost": round(float(valid["transport_cost"].sum()), 2),
            "total_procurement_cost": round(float(valid["procurement_cost"].sum()), 2),
            "total_holding_cost": round(float(valid["holding_cost"].sum()), 2),
            "total_shortage_cost": round(float(valid["shortage_cost"].sum()), 2),
            "avg_transport_cost": round(float(valid["transport_cost"].mean()), 2),
            "avg_shortage_cost": round(float(valid["shortage_cost"].mean()), 2),
            "avg_holding_cost": round(float(valid["holding_cost"].mean()), 2),
            "avg_procurement_cost": round(float(valid["procurement_cost"].mean()), 2),
            "periods_solved": len(valid),
            "periods_failed": len(metrics_df) - len(valid),
        }

    return OptimizationResult(
        region=req.region,
        strategy=req.strategy,
        periods=periods,
        summary=summary,
    )


@router.post("/plan")
def run_plan(req: PlanningRequest):
    try:
        if req.use_cms_data:
            instance = build_cms_region_instance(req.region, scenario=req.scenario)
        else:
            instance = build_region_instance(req.region)
    except ValueError as e:
        raise HTTPException(400, str(e))

    if req.arc_cap is not None:
        instance["arc_cap"] = float(req.arc_cap)

    # Apply custom prices if provided
    proc_cost = instance.get("proc_cost", 0.0)
    if req.custom_prices and not np.isscalar(proc_cost):
        proc_cost = pd.Series(proc_cost).copy()
        for drug, price in req.custom_prices.items():
            if drug in proc_cost.index:
                proc_cost[drug] = float(price)

    # Shortage penalty = multiplier × per-drug procurement cost
    if np.isscalar(proc_cost):
        shortage_pen = req.shortage_penalty * float(proc_cost) if proc_cost else 10.0
    else:
        shortage_pen = pd.Series(proc_cost).clip(lower=1.0) * req.shortage_penalty

    try:
        result = run_planning(
            instance=instance,
            strategy=req.strategy,
            kappa=req.kappa,
            Gamma=req.gamma,
            transport_cost_per_km=req.transport_cost_per_km,
            shortage_penalty=shortage_pen,
            holding_cost=req.holding_cost,
            procurement_cost=proc_cost,
            initial_inventory=req.initial_inventory,
            last_demand=req.last_demand,
        )
    except Exception as e:
        raise HTTPException(500, f"Planning failed: {e}")

    return {"region": req.region, **result}


def _build_region_template(region: str) -> io.BytesIO:
    """Build an xlsx workbook with Inventory and Order-Request sheets for a region."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        instance = build_cms_region_instance(region, scenario="2526")
    except ValueError as e:
        raise HTTPException(400, str(e))

    facilities = list(instance["mu_mat"].index)
    drugs = list(instance["mu_mat"].columns)

    cms = app_data.cms_active
    desc_lookup = {code: str(cms.loc[code].get("description", "")) for code in drugs if code in cms.index}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1f4e79")
    subheader_font = Font(italic=True, color="555555")

    def build_sheet(ws, title: str, instructions: str):
        ws.title = title
        ws["A1"] = f"{title} - {region}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = instructions
        ws["A2"].font = subheader_font
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=min(8, len(drugs) + 1))
        ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=min(8, len(drugs) + 1))

        header_row = 4
        ws.cell(row=header_row, column=1, value="Facility").font = header_font
        ws.cell(row=header_row, column=1).fill = header_fill
        for j, drug in enumerate(drugs, start=2):
            cell = ws.cell(row=header_row, column=j, value=drug)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Drug description sub-header
        desc_row = header_row + 1
        ws.cell(row=desc_row, column=1, value="").font = subheader_font
        for j, drug in enumerate(drugs, start=2):
            desc = desc_lookup.get(drug, "")
            c = ws.cell(row=desc_row, column=j, value=desc)
            c.font = subheader_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)

        # Facility rows
        for i, fac in enumerate(facilities, start=desc_row + 1):
            ws.cell(row=i, column=1, value=fac).font = Font(bold=True)
            for j in range(2, len(drugs) + 2):
                ws.cell(row=i, column=j, value=None)

        # Column widths
        ws.column_dimensions["A"].width = 40
        for j in range(2, len(drugs) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 14
        ws.freeze_panes = ws.cell(row=desc_row + 1, column=2)

    build_sheet(
        wb.active,
        "Inventory",
        "Enter current stock on hand per facility and drug. Blank cells will be treated as 0.",
    )
    order_ws = wb.create_sheet()
    build_sheet(
        order_ws,
        "OrderRequest",
        "Enter this cycle's actual order request per facility and drug. Blank cells will be treated as 0.",
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/regions/{region}/template")
def download_region_template(region: str):
    buf = _build_region_template(region)
    safe = region.replace(" ", "_").replace("/", "-")
    headers = {
        "Content-Disposition": f'attachment; filename="{safe}_template.xlsx"',
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/plan/export")
def export_plan(body: dict):
    """
    Export a plan result (from /plan) to a formatted xlsx workbook.
    Body: { region, strategy, summary, shipments, procurement, round_up: bool }
    """
    import math
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    region = body.get("region", "Region")
    strategy = body.get("strategy", "")
    summary = dict(body.get("summary", {}) or {})
    shipments = list(body.get("shipments", []) or [])
    procurement = list(body.get("procurement", []) or [])
    round_up = bool(body.get("round_up", False))

    # Optional rounding: ceil quantities and recompute totals.
    if round_up:
        for s in shipments:
            s["quantity"] = float(math.ceil(float(s.get("quantity", 0))))
        total_procurement_cost = 0.0
        total_units_procured = 0.0
        for p in procurement:
            qty = float(math.ceil(float(p.get("quantity", 0))))
            unit = float(p.get("unit_cost", 0))
            p["quantity"] = qty
            p["total_cost"] = round(qty * unit, 2)
            total_procurement_cost += qty * unit
            total_units_procured += qty
        summary["total_procurement_cost"] = round(total_procurement_cost, 2)
        summary["total_units_procured"] = round(total_units_procured, 1)
        summary["total_units_shipped"] = round(
            sum(float(s.get("quantity", 0)) for s in shipments), 1
        )
        summary["rounding"] = "whole units (ceil)"

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1f4e79")
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Shipment Plan - {region}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Strategy: {strategy}"
    ws["A2"].font = Font(italic=True, color="555555")
    if round_up:
        ws["A3"] = "Quantities rounded up to whole units; procurement cost recomputed."
        ws["A3"].font = Font(italic=True, color="8b5a00")

    metric_labels = [
        ("active_routes", "Active Routes"),
        ("total_shipments", "Total Shipment Lines"),
        ("total_units_shipped", "Total Units to Ship"),
        ("total_procurement_orders", "Procurement Lines"),
        ("total_units_procured", "Total Units to Procure"),
        ("total_procurement_cost", "Total Procurement Cost (BWP)"),
        ("total_transport_cost", "Total Transport Cost (BWP)"),
        ("total_holding_cost", "Realized Holding Cost (BWP)"),
        ("total_shortage_cost", "Realized Shortage Cost (BWP)"),
    ]
    r = 5
    for key, label in metric_labels:
        v = summary.get(key)
        if v is None:
            continue
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=v)
        r += 1
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20

    # Shipments sheet
    ship_ws = wb.create_sheet("Shipments")
    ship_headers = ["From", "To", "Drug", "Quantity", "Distance (km)"]
    for j, h in enumerate(ship_headers, start=1):
        c = ship_ws.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for i, s in enumerate(shipments, start=2):
        ship_ws.cell(row=i, column=1, value=s.get("from", ""))
        ship_ws.cell(row=i, column=2, value=s.get("to", ""))
        ship_ws.cell(row=i, column=3, value=s.get("drug", ""))
        ship_ws.cell(row=i, column=4, value=s.get("quantity", 0))
        ship_ws.cell(row=i, column=5, value=s.get("distance_km", 0))
        for j in range(1, 6):
            ship_ws.cell(row=i, column=j).border = border
    for col, w in zip("ABCDE", [28, 28, 12, 14, 14]):
        ship_ws.column_dimensions[col].width = w
    ship_ws.freeze_panes = "A2"
    ship_ws.auto_filter.ref = ship_ws.dimensions

    # Procurement sheet
    proc_ws = wb.create_sheet("Procurement")
    proc_headers = ["Drug", "Quantity", "Unit Cost (BWP)", "Total Cost (BWP)"]
    for j, h in enumerate(proc_headers, start=1):
        c = proc_ws.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for i, p in enumerate(procurement, start=2):
        proc_ws.cell(row=i, column=1, value=p.get("drug", ""))
        proc_ws.cell(row=i, column=2, value=p.get("quantity", 0))
        proc_ws.cell(row=i, column=3, value=p.get("unit_cost", 0))
        proc_ws.cell(row=i, column=4, value=p.get("total_cost", 0))
        for j in range(1, 5):
            proc_ws.cell(row=i, column=j).border = border
    for col, w in zip("ABCD", [14, 14, 18, 18]):
        proc_ws.column_dimensions[col].width = w
    proc_ws.freeze_panes = "A2"
    proc_ws.auto_filter.ref = proc_ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = region.replace(" ", "_").replace("/", "-")
    suffix = "_rounded" if round_up else ""
    headers = {
        "Content-Disposition": f'attachment; filename="{safe}_plan{suffix}.xlsx"',
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/regions/{region}/parse-template")
async def parse_region_template(region: str, file: UploadFile = File(...)):
    """Parse an uploaded filled template and return {inventory, order_request} dicts."""
    from openpyxl import load_workbook

    try:
        instance = build_cms_region_instance(region, scenario="2526")
    except ValueError as e:
        raise HTTPException(400, str(e))

    valid_facilities = set(instance["mu_mat"].index)
    valid_drugs = list(instance["mu_mat"].columns)

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read workbook: {e}")

    def parse_sheet(ws) -> dict:
        # Header on row 4, descriptions row 5, data starts row 6
        header = [ws.cell(row=4, column=c).value for c in range(2, ws.max_column + 1)]
        drug_cols = {c + 2: str(d).strip() for c, d in enumerate(header) if d}
        result: dict[str, dict[str, float]] = {}
        for r in range(6, ws.max_row + 1):
            fac = ws.cell(row=r, column=1).value
            if not fac:
                continue
            fac = str(fac).strip()
            if fac not in valid_facilities:
                continue
            row_vals: dict[str, float] = {}
            for col, drug in drug_cols.items():
                if drug not in valid_drugs:
                    continue
                v = ws.cell(row=r, column=col).value
                if v is None or v == "":
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                if fv != 0:
                    row_vals[drug] = fv
            if row_vals:
                result[fac] = row_vals
        return result

    inventory = parse_sheet(wb["Inventory"]) if "Inventory" in wb.sheetnames else {}
    order_request = parse_sheet(wb["OrderRequest"]) if "OrderRequest" in wb.sheetnames else {}

    return {
        "region": region,
        "inventory": inventory,
        "order_request": order_request,
        "inventory_rows": len(inventory),
        "order_request_rows": len(order_request),
    }
